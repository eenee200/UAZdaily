import json
from datetime import datetime, timedelta
import tempfile
from bs4 import BeautifulSoup
from openpyxl.utils.dataframe import dataframe_to_rows
import re
import pandas as pd
import numpy as np
from openpyxl import Workbook
from openpyxl.styles import Border, Side, Alignment,Font,PatternFill

def parse_data(raw_data, engine_raw_data=None):
    """
    Parse fuel data and filter by engine status if available
    """
    data_str = raw_data.strip()
    data_points = [point.strip('[]').split(',') for point in data_str.split('],[')]
    valid_fuel = None
    
    # Parse engine data if provided, handling duplicate timestamps
    engine_status = {}
    engine_history = {}  # To track engine status history
    last_engine_status = 1  # Default to on if no engine data
    
    if engine_raw_data:
        engine_str = engine_raw_data.strip()
        engine_points = [point.strip('[]').split(',') for point in engine_str.split('],[')]
        
        # Sort by timestamp to process in chronological order
        try:
            engine_points.sort(key=lambda x: float(x[0]))
        except (ValueError, IndexError):
            print("Warning: Could not sort engine points")
        
        for point in engine_points:
            if len(point) >= 2:
                try:
                    timestamp = float(point[0])
                    status = int(float(point[1].strip('"')))  # 1 = on, 0 = off
                    
                    # Track the history of engine status
                    if timestamp not in engine_history:
                        engine_history[timestamp] = []
                    engine_history[timestamp].append(status)
                    
                    # For the same timestamp, if ANY status is 1 (on), consider the engine on
                    if timestamp in engine_status:
                        engine_status[timestamp] = max(engine_status[timestamp], status)
                    else:
                        engine_status[timestamp] = status
                    
                    last_engine_status = status
                except (ValueError, IndexError) as e:
                    print(f"Warning: Could not parse engine point: {point}, Error: {str(e)}")
    
    # Handle duplicated timestamps in engine data
    for timestamp in engine_history:
        if 1 in engine_history[timestamp]:
            engine_status[timestamp] = 1  # If engine was on at any point at this timestamp, consider it on
    
    parsed_data = []
    fuel_values = set()
    current_engine_state = last_engine_status  # Start with the last known engine state
    
    # Sort data points by timestamp
    try:
        data_points.sort(key=lambda x: float(x[0]))
    except (ValueError, IndexError):
        print("Warning: Could not sort fuel data points")
    
    for point in data_points:
        if len(point) >= 2:
            try:
                timestamp = float(point[0])
                fuel = float(point[1].strip('"'))
                
                # Update engine state if we have a reading at this timestamp
                if timestamp in engine_status:
                    current_engine_state = engine_status[timestamp]
                
                # Only include fuel readings when engine is on (1)
                if engine_raw_data is None or current_engine_state == 1:
                    if fuel == 0 and valid_fuel is not None:
                        fuel = valid_fuel
                    
                    if fuel != 0:
                        valid_fuel = fuel
                        parsed_data.append((timestamp, fuel))
                        fuel_values.add(fuel)
            except (ValueError, IndexError) as e:
                print(f"Warning: Could not parse fuel point: {point}, Error: {str(e)}")
    
    # Check if all fuel values are the same or if we have no valid data
    if len(fuel_values) <= 1:
        return None  # Return None if all fuel values are the same or empty
    else:
        return parsed_data

def load_data_from_file(file_path, engine_file=None):
    """
    Load data from HTML file with optional engine status filtering
    """
    with open(file_path, 'r', encoding='utf-8') as file:
        content = file.read()

    # Load engine file if provided
    engine_content = None
    if engine_file:
        try:
            with open(engine_file, 'r', encoding='utf-8') as eng_file:
                engine_content = eng_file.read()
        except Exception as e:
            print(f"Warning: Failed to load engine status data: {str(e)}")

    soup = BeautifulSoup(content, 'html.parser')
    engine_soup = BeautifulSoup(engine_content, 'html.parser') if engine_content else None
    
    datasets = []
    engine_datasets = []
    identifiers = []
    date_range = None
    identifiers_to_remove = set()

    # Find all script tags and object rows in fuel data
    all_elements = soup.find_all(['script', 'table'])
    current_identifier = None
    current_datasets = []

    # Extract engine data if available
    engine_data_by_identifier = {}
    if engine_soup:
        all_engine_elements = engine_soup.find_all(['script', 'table'])
        current_engine_identifier = None
        current_engine_datasets = []
        
        for element in all_engine_elements:
            if element.name == 'table':
                # Process previous identifier's datasets
                if current_engine_identifier and current_engine_datasets:
                    engine_data_by_identifier[current_engine_identifier] = current_engine_datasets
                
                # Get new identifier
                object_row = element.find('td', string='Обьект:')
                if object_row:
                    current_engine_identifier = object_row.find_next_sibling('td').get_text(strip=True)
                    current_engine_datasets = []
            
            elif element.name == 'script' and element.string:
                # Collect engine data matches
                data_matches = re.findall(r'data":\s*(\[.*?\])\s*,\s*"data_index', element.string, re.DOTALL)
                current_engine_datasets.extend(data_matches)
        
        # Process last identifier's datasets
        if current_engine_identifier and current_engine_datasets:
            engine_data_by_identifier[current_engine_identifier] = current_engine_datasets

    # Process fuel data
    for element in all_elements:
        if element.name == 'table':
            # Check for object identifier
            object_row = element.find('td', string='Обьект:')
            if object_row:
                # Process previous identifier's datasets if exists
                if current_identifier and current_datasets:
                    valid_datasets = []
                    for i, dataset in enumerate(current_datasets):
                        # Get corresponding engine data if available
                        engine_data = None
                        if current_identifier in engine_data_by_identifier and i < len(engine_data_by_identifier[current_identifier]):
                            engine_data = engine_data_by_identifier[current_identifier][i]
                        
                        parsed_data = parse_data(dataset, engine_data)
                        if parsed_data is not None:
                            valid_datasets.append((dataset, engine_data))
                    
                    # Add to identifiers and datasets, or mark for removal
                    if valid_datasets:
                        # Handle multiple valid datasets for same identifier
                        if len(valid_datasets) > 1:
                            datasets.append(valid_datasets[0])
                            identifiers.append(current_identifier + " 1")
                            
                            if len(valid_datasets) >= 2:
                                datasets.append(valid_datasets[1])
                                identifiers.append(current_identifier + " 2")
                                if len(valid_datasets) >= 3:
                                    datasets.append(valid_datasets[2])
                                    identifiers.append(current_identifier + " 3")
                        else:
                            # Single dataset case
                            datasets.extend(valid_datasets)
                            identifiers.append(current_identifier)
                    else:
                        identifiers_to_remove.add(current_identifier)

                # Reset for new identifier
                current_identifier = object_row.find_next_sibling('td').get_text(strip=True)
                current_datasets = []

                # Check for date range
                date_row = element.find('td', string='Хугацаа:')
                if date_row:
                    date_cell = date_row.find_next_sibling('td')
                    if date_cell and not date_range:
                        date_range = date_cell.get_text(strip=True)

        elif element.name == 'script' and element.string:
            # Collect fuel data matches
            data_matches = re.findall(r'data":\s*(\[.*?\])\s*,\s*"data_index', element.string, re.DOTALL)
            current_datasets.extend(data_matches)

    # Process last identifier's datasets
    if current_identifier and current_datasets:
        valid_datasets = []
        engine_data = None
        for i, dataset in enumerate(current_datasets):
            # Get corresponding engine data if available
            
            if current_identifier in engine_data_by_identifier and i < len(engine_data_by_identifier[current_identifier]):
                engine_data = engine_data_by_identifier[current_identifier][i]
            
            parsed_data = parse_data(dataset, engine_data)
            if parsed_data is not None:
                valid_datasets.append((dataset, engine_data))
        
        if valid_datasets:
            # Handle multiple valid datasets for same identifier
            if len(valid_datasets) > 1:
                datasets.append(valid_datasets[0])
                identifiers.append(current_identifier + " 1")
                
                if len(valid_datasets) >= 2:
                    datasets.append(valid_datasets[1])
                    identifiers.append(current_identifier + " 2")
            else:
                datasets.extend(valid_datasets)
                identifiers.append(current_identifier)
        else:
            identifiers_to_remove.add(current_identifier)

    # Remove identifiers with no valid data
    final_identifiers = [ident for ident in identifiers if ident not in identifiers_to_remove]
    
    if not datasets:
        raise ValueError("No data arrays found in the file.")

    return datasets, final_identifiers, [date_range] if date_range else ['']


def load_daily_distances(file_path, valid_identifiers):
    with open(file_path, 'r', encoding='utf-8') as file:
        content = file.read()

    soup = BeautifulSoup(content, 'html.parser')
    all_daily_distances = []
    all_daily_dates = []
    found_identifiers = []
    deleted_distances = []
    deleted_dates = []
    deleted_identifiers = []

    object_rows = soup.find_all('td', string='Обьект:')
    
    for object_row in object_rows:
        identifier = object_row.find_next_sibling('td').get_text(strip=True)
        current_table = object_row.find_parent('table')
        distance_table = current_table.find_next_sibling('table')
        
        if distance_table:
            daily_distances = []
            daily_dates = []
            for row in distance_table.find_all('tr')[1:]:
                cells = row.find_all('td')
                if len(cells) >= 2:
                    date = datetime.strptime(cells[0].get_text(strip=True), '%Y-%m-%d').date()
                    distance = cells[1].get_text(strip=True)
                    distance_val = float(distance.replace(' km', ''))
                    daily_distances.append(distance_val)
                    daily_dates.append(date)
            
            # Modified: Check for base identifier (without " 1" or " 2" suffix)
            base_identifier_found = False
            for valid_id in valid_identifiers:
                # Strip any " 1" or " 2" suffix for comparison
                base_valid_id = valid_id.split(" ")[0] if " " in valid_id else valid_id
                
                if identifier == base_valid_id:
                    base_identifier_found = True
                    found_identifiers.append(identifier)
                    all_daily_distances.append(daily_distances)
                    all_daily_dates.append(daily_dates)
                    break
            
            if not base_identifier_found:
                deleted_identifiers.append(identifier)
                deleted_distances.append(daily_distances)
                deleted_dates.append(daily_dates)

    # Modified: Map the daily distances to both identifier variations (with " 1" and " 2" suffixes)
    ordered_distances = []
    ordered_dates = []
    
    for valid_id in valid_identifiers:
        # Check if this is a derived identifier (with " 1" or " 2" suffix)
        base_id = valid_id.split(" ")[0] if " " in valid_id else valid_id
        
        # Find the base identifier in found_identifiers
        if base_id in found_identifiers:
            idx = found_identifiers.index(base_id)
            ordered_distances.append(all_daily_distances[idx])
            ordered_dates.append(all_daily_dates[idx])
        else:
            # If we can't find the base identifier, check for exact match
            if valid_id in found_identifiers:
                idx = found_identifiers.index(valid_id)
                ordered_distances.append(all_daily_distances[idx])
                ordered_dates.append(all_daily_dates[idx])

    final_distances = ordered_distances + deleted_distances
    final_dates = ordered_dates + deleted_dates
    final_identifiers = valid_identifiers + deleted_identifiers

    return final_distances, final_identifiers, final_dates






def detect_refills(data, threshold_percentage=5, time_window_minutes=60):
    refills = []
    in_refill = False
    min_fuel = None
    max_fuel = None
    start_time = None
    last_valid_fuel = None  # To store the last fuel value greater than 3

    def check_previous_fuel_levels(data, current_index, start_time, max_fuel):
        """Check if there's a higher fuel level in the previous 120 minutes"""
        check_start_time = start_time - timedelta(minutes=120)
        comparison_fuel = max_fuel - 5
        check_end_time = start_time
        for j in range(current_index - 1, -1, -1):
            check_time = datetime.utcfromtimestamp(data[j][0]/1000)
            if check_time < check_start_time:
                break
            if check_time > check_end_time:
                continue
            if data[j][1] > comparison_fuel:
                return True
        return False
    
    def find_real_start_time(data, start_idx, start_fuel):
        """Find the actual start time by skipping over periods of constant fuel level"""
        real_start_idx = start_idx
        current_fuel = start_fuel
        
        for i in range(start_idx + 1, len(data)):
            if data[i][1] > current_fuel:
                real_start_idx = i - 1
                break
            if data[i][1] < current_fuel:
                break
        
        return datetime.utcfromtimestamp(data[real_start_idx][0]/1000)
    
    def check_data_exists_in_window(data, check_time, current_index):
        
        # if current_index < 0 or current_index >= len(data):
        #     return False, False, False
            
        current_time = check_time
        
        # Define window boundaries
        time_5min_before = current_time - timedelta(minutes=10)
        time_30sec_after = time_5min_before + timedelta(seconds=30)
        time_30sec_before = time_5min_before - timedelta(seconds=30)
        
        exists_before_5min = False
        
        # Check if any data exists before the 5-minute boundary
        for j in range(current_index - 1, -1, -1):
            check_time = datetime.utcfromtimestamp(data[j][0]/1000)
            if time_30sec_before < check_time < time_30sec_after:
                exists_before_5min = True
                break
            if check_time < time_5min_before:
                break
        
       
        
        return exists_before_5min
        
    for i in range(1, len(data)):
        prev_fuel = data[i-1][1]
        current_fuel = data[i][1]
        current_time = datetime.utcfromtimestamp(data[i][0]/1000)

        if current_fuel == None:
            current_fuel = 0
        if prev_fuel == None:
            prev_fuel = 0
        
        if current_fuel >= 1:
            last_valid_fuel = current_fuel
        
        if current_fuel >= prev_fuel:
            if not in_refill:
                # Start a refill only if there's data before our window
                start_time = find_real_start_time(data, i-1, prev_fuel)
                if check_data_exists_in_window(data, start_time, i-1):
                    in_refill = True
                    min_fuel = prev_fuel if prev_fuel >= 1 else last_valid_fuel
                    start_time = find_real_start_time(data, i-1, prev_fuel)
            if in_refill:
                max_fuel = current_fuel
                last_time = current_time
        elif in_refill:
            in_refill = False
            if min_fuel is not None and max_fuel is not None:
                if min_fuel <= 0 and last_valid_fuel is not None:
                    min_fuel = last_valid_fuel
                
                percent_change = max_fuel - min_fuel
                if min_fuel >= 0:
                    if percent_change > threshold_percentage:
                        valid_refill = True
                        end_time = last_time + timedelta(minutes=time_window_minutes)
                        
                        if check_previous_fuel_levels(data, i, start_time, max_fuel):
                            valid_refill = False
                        else:
                            # Check for significant drops after the refill
                            for j in range(i, len(data)):
                                check_time = datetime.utcfromtimestamp(data[j][0]/1000)
                                if check_time > end_time:
                                    break
                                # Only invalidate if we see a significant drop
                                if data[j][1] <= min_fuel + (percent_change * 0.7):  # Allow for some normal usage drop
                                    valid_refill = False
                                    break
                        
                        if valid_refill:
                            if refills and (last_time - refills[-1]['timestamp']) <= timedelta(minutes=time_window_minutes):
                                refills[-1]['max_fuel'] = max(refills[-1]['max_fuel'], max_fuel)
                                refills[-1]['percent_change'] = refills[-1]['max_fuel'] - refills[-1]['min_fuel']
                            else:
                                refills.append({
                                    'timestamp': start_time,
                                    'percent_change': percent_change,
                                    'max_fuel': max_fuel,
                                    'min_fuel': min_fuel
                                })
            min_fuel, max_fuel = None, None
    
    return refills
def analyze_fuel_data(data_pair):
    """
    Analyze fuel data with engine status filtering
    """
    raw_data, engine_data = data_pair
    data = parse_data(raw_data, engine_data)
    
    if not data:
        return [], {'num_refills': 0, 'first_fuel': None, 'last_fuel': None}
    
    first_fuel = data[0][1]
    last_fuel = data[-1][1]
    refills = detect_refills(data)
    
    stats = {
        'num_refills': len(refills),
        'first_fuel': first_fuel,
        'last_fuel': last_fuel,
    }
    
    return refills, stats

def export_to_excel(datasets, identifiers, date_ranges, all_daily_distances, all_daily_dates, output_file='fuel_analysis.xlsx'):
    try:
        all_summary_data = []
        all_refills_data = []
        all_daily_data = []
        urgent_check_needed = []
        
        # Process multiple datasets
        for idx, (refills, stats, data) in enumerate(datasets):
            dataset_name = identifiers[idx] 
            refills_data = []
            total_refill = 0.0  # Initialize as float
            total_consumption = 0.0  # Initialize as float
            
            # Safely handle None values for first and last fuel readings
            first = float(stats['first_fuel'] if stats['first_fuel'] is not None else 0)
            last = float(stats['last_fuel'] if stats['last_fuel'] is not None else 0)
            daily_distances = all_daily_distances[idx] if idx < len(all_daily_distances) else []

            # Create a dictionary to store refill dates for daily counting
            refill_dates = {}
            for refill in refills:
                refill_date = refill['timestamp'].date()
                refill_dates[refill_date] = refill_dates.get(refill_date, 0) + 1
            
            for i, refill in enumerate(refills, 1):
                # Safely handle None values in refill calculations
                min_fuel = float(refill.get('min_fuel', 0) or 0)  # Convert None to 0
                max_fuel = float(refill.get('max_fuel', 0) or 0)  # Convert None to 0
                
                consumption = round(first - min_fuel, 2)
                percent_change = max_fuel - min_fuel
                
                total_refill += percent_change
                total_consumption += consumption

                refills_data.append({
                    ' ': " ",
                    'Эхэлсэн хугацаа': refill['timestamp'],
                    'Өмнөх түлш': round(min_fuel, 2),
                    'Дараах түлш': round(max_fuel, 2),
                    'Нэмсэн түлш': round(percent_change, 2),
                    'Сүүлд дүүргэснээс хойш зарцуулалт': round(consumption, 2)
                })

                first = max_fuel

            # Safely calculate final consumption
            final_consumption = first - last if first is not None and last is not None else 0
            total_consumption += final_consumption

            # Calculate distance per refill safely
            total_distance = sum(daily_distances) if daily_distances else 0
            avg_consumption = (total_consumption / total_distance * 100) if total_distance > 0 else 0

            # Create summary data with safe handling of None values
            summary_data = {
                'Обьект': dataset_name + (" (яаралтай шалгуулах хэрэгтэй)" if idx in urgent_check_needed else ""),
                'Нийт явсан км': total_distance if total_distance != 0 else 'N/A',
                'Түлш дүүрлт /Л/': round(float(total_refill), 2),
                'Түлш дүүргэсэн тоо': int(stats.get('num_refills', 0)),
                'Түлш зарцуулалт /Л/': round(float(total_consumption), 2) if total_consumption > 0 else 0,
                'Дундаж хэрэглээ/100км/': round(avg_consumption, 2) if avg_consumption > 0 else "",
                'Эхний үлдэгдэл': round(float(stats.get('first_fuel', 0) or 0), 2),
                'Эцсийн үлдэгдэл': round(float(stats.get('last_fuel', 0) or 0), 2)
            }
            all_summary_data.append(summary_data)

            # Check for urgent cases
            if stats.get('first_fuel', 0) == 0 and stats.get('last_fuel', 0) == 0:
                urgent_check_needed.append(idx)

            # Check for multiple refills in 24 hours
            refill_times = [refill['timestamp'] for refill in refills]
            refill_times.sort()
            for i in range(len(refill_times)):
                end_time = refill_times[i]
                start_time = end_time - timedelta(hours=24)
                refills_in_24h = sum(1 for t in refill_times if start_time <= t <= end_time)
                if refills_in_24h >= 5:
                    urgent_check_needed.append(idx)
                    break

            # Process daily data
            daily_data = []
            daily_dates = all_daily_dates[idx] if idx < len(all_daily_dates) else []
            daily_start_fuel = 0.0
            daily_end_fuel = 0.0
            
            for date_idx, current_date in enumerate(daily_dates):
                # Find fuel levels for this date
                day_fuel_levels = []
                total_daily_refill = 0.0
                
                for timestamp, fuel_level in data:
                    data_date = datetime.utcfromtimestamp(timestamp / 1000).date()
                    if data_date == current_date:
                        day_fuel_levels.append(float(fuel_level if fuel_level is not None else 0))
                        
                # Get start and end fuel levels for the day
                if day_fuel_levels:
                    daily_start_fuel = day_fuel_levels[0]
                    daily_end_fuel = day_fuel_levels[-1]
                
                # Calculate total daily refill amount
                for refill in refills:
                    if refill['timestamp'].date() == current_date:
                        total_daily_refill += float(refill.get('percent_change', 0) or 0)
                
                daily_consumption = daily_start_fuel + total_daily_refill - daily_end_fuel
                daily_distance = float(daily_distances[date_idx] if date_idx < len(daily_distances) else 0)
                
                # Calculate average consumption per 100km
                avg_consumption = (daily_consumption / daily_distance * 100) if daily_distance > 0 else 0
                
                daily_data.append({
                    '': current_date,
                    'Нийт явсан км': round(daily_distance, 2),
                    'Түлш дүүрлт /Л/': round(total_daily_refill, 2),
                    'Түлш дүүргэсэн тоо': refill_dates.get(current_date, 0),
                    'Түлш зарцуулалт /Л/': round(daily_consumption, 2) if daily_consumption > 0 else 0,
                    'Дундаж хэрэглээ/100км/': round(avg_consumption, 2) if avg_consumption > 0 else " ",
                    'Эхний үлдэгдэл': round(daily_start_fuel, 2),
                    'Эцсийн үлдэгдэл': round(daily_end_fuel, 2),
                })
            
            all_refills_data.append((dataset_name, refills_data))
            all_daily_data.append((dataset_name, daily_data))

        # Export all data to Excel
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = 'Ерөнхий мэдээлэл'
        
        date_range_cell = worksheet.cell(row=1, column=1, value=f"Хугацаа: {date_ranges[0]}")
        date_range_cell.font = Font(bold=True)

        # Write summary data
        summary_df = pd.DataFrame(all_summary_data)
        header_row = dataframe_to_rows(summary_df, index=False, header=True)
        worksheet.append(next(header_row))
        
       
        # Add datasets to the Excel file
        for idx, (dataset_name, refills_data) in enumerate(all_refills_data):
               # Write object name (identifier) and date range only once at the top
            # Write summary for each dataset
            daily_distances = all_daily_distances[idx] if idx < len(all_daily_distances) else []
            summary_start_row = worksheet.max_row   # Leave a gap before the next dataset
            total_distance = sum(daily_distances) if daily_distances else 0
            total_refill = all_summary_data[idx]['Түлш зарцуулалт /Л/']

    # Calculate distance per refill, handle division by zero if total_refill is 0
            distance_per_refill = total_refill / total_distance if total_distance != 0 else 'N/A'
            distance_per_refill = distance_per_refill * 100
            summary_row = {
                'Обьект': dataset_name + (" (яаралтай шалгуулах хэрэгтэй)" if idx in urgent_check_needed else ""),
                'Нийт явсан км': sum(daily_distances) if daily_distances else 'N/A',
                'Түлш дүүрлт /Л/': round(all_summary_data[idx]['Түлш дүүрлт /Л/'], 2),
                'Түлш дүүргэсэн тоо': all_summary_data[idx]['Түлш дүүргэсэн тоо'],
                'Түлш зарцуулалт /Л/': round(all_summary_data[idx]['Түлш зарцуулалт /Л/'], 2),
                'Дундаж хэрэглээ/100км/': round(distance_per_refill, 2) if isinstance(distance_per_refill, (int, float)) else 'N/A',
                'Эхний үлдэгдэл': round(all_summary_data[idx]['Эхний үлдэгдэл'], 2),
                'Эцсийн үлдэгдэл': round(all_summary_data[idx]['Эцсийн үлдэгдэл'], 2)
            }
            worksheet.append([summary_row[key] for key in summary_row])
            if idx in urgent_check_needed:
                cell = worksheet.cell(row=worksheet.max_row, column=1)
                cell.fill = PatternFill(start_color="FFCCCB", end_color="FFCCCB", fill_type="solid")  # Light red background
            else:
                cell = worksheet.cell(row=worksheet.max_row, column=1)
                cell.fill = PatternFill(start_color="D8E4BC", end_color="D8E4BC", fill_type="solid")  
            fill_color = PatternFill(start_color="B8CCE4", end_color="B8CCE4", fill_type="solid")
            for col in range(1, 9):  # Adjust the range for the number of columns you want
                cell = worksheet.cell(row=2, column=col)
                cell.fill = fill_color
            # Write refills data
            refills_df = pd.DataFrame(refills_data)
            daily_row = worksheet.max_row + 1
            for r in dataframe_to_rows(refills_df, index=False, header=True):
                worksheet.append(r)
            fil_color = PatternFill(start_color="E4DFEC", end_color="E4DFEC", fill_type="solid")
            for col in range(2, 9):  # Adjust the range for the number of columns you want
                cell = worksheet.cell(row=daily_row, column=col)
                cell.fill = fil_color

            daily_df = pd.DataFrame(all_daily_data[idx][1])
            daily_start_row = worksheet.max_row + 1
            
            for r in dataframe_to_rows(daily_df, index=False, header=True):
                worksheet.append(r)
            
            # Get the starting and ending rows for daily data
            daily_end_row = worksheet.max_row
            worksheet.row_dimensions.group(daily_start_row , daily_end_row, outline_level=1, hidden=True)
            for col in range(2, 9):  # Adjust the range for the number of columns you want
                cell = worksheet.cell(row=daily_start_row, column=col)
                cell.fill = fil_color
            

        # Apply formatting
        thin_border = Border(left=Side(style='thin'),
                             right=Side(style='thin'),
                             top=Side(style='thin'),
                             bottom=Side(style='thin'))
        
        # Apply border and alignment to all cells
        for row in worksheet.iter_rows(min_row=1, max_row=worksheet.max_row, min_col=1, max_col=worksheet.max_column):
            for cell in row:
                cell.border = thin_border
                cell.alignment = Alignment(wrap_text=True, vertical='center', horizontal='center')

        # Adjust column widths
        for column in worksheet.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    max_length = max(max_length, len(str(cell.value)))
                except:
                    pass
            adjusted_width = (max_length + 2)
            worksheet.column_dimensions[column_letter].width = adjusted_width

        # Save the workbook
        workbook.save(output_file)

        return output_file, len(datasets)

    except Exception as e:
        print(f"Error exporting to Excel: {str(e)}")
        return None, 0


def main(file_path1, file_path2, engine_file=None):
    all_datasets = []
    all_identifiers = []
    all_date_ranges = []

    # Load datasets from the first HTML file with engine status
    raw_datasets, active_identifiers, date_ranges = load_data_from_file(file_path1, engine_file)
    all_identifiers.extend(active_identifiers)
    all_date_ranges.extend(date_ranges)

    # Process each dataset from file_path1
    for idx, data_pair in enumerate(raw_datasets):
        raw_data, engine_data = data_pair
        refills, stats = analyze_fuel_data(data_pair)
        data = parse_data(raw_data, engine_data)
        all_datasets.append((refills, stats, data))

    # Load all daily distances and dates
    all_daily_distances, combined_identifiers, all_daily_dates = load_daily_distances(file_path2, active_identifiers)
    
    # Find removed identifiers
    removed_identifiers = [id for id in combined_identifiers if id not in active_identifiers]
    all_identifiers = active_identifiers + removed_identifiers

    # Create empty datasets for removed identifiers
    for i, removed_id in enumerate(removed_identifiers):
        # Find the index in combined_identifiers
        try:
            idx = combined_identifiers.index(removed_id)
            # Check if the index is valid for all_daily_dates
            if idx < len(all_daily_dates):
                daily_dates = all_daily_dates[idx]
            else:
                daily_dates = []
        except (ValueError, IndexError):
            # Handle case where index is not found or out of range
            daily_dates = []
        
        empty_refills = []
        empty_stats = {'num_refills': 0, 'first_fuel': 0, 'last_fuel': 0}
        empty_data = []

        if daily_dates:
            # Create data points for each day
            for date in daily_dates:
                timestamp = int(datetime.combine(date, datetime.min.time()).timestamp() * 1000)
                empty_data.append((timestamp, 0))
        
        all_datasets.append((empty_refills, empty_stats, empty_data))

    # Create temporary file for Excel output
    with tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx') as tmp:
        temp_path = tmp.name

    # Export data to Excel
    excel_file, num_datasets = export_to_excel(
        all_datasets, 
        all_identifiers, 
        all_date_ranges, 
        all_daily_distances,
        all_daily_dates,
        output_file=temp_path
    )

    return excel_file, num_datasets

if __name__ == "__main__":
    file_path1 = 'C:/Users/User/Desktop/ttt/web/test/tulsh.html'
    file_path2 = 'C:/Users/User/Desktop/ttt/web/test/zam.html'
    engine_file = 'C:/Users/User/Desktop/ttt/web/test/tog.html'  # Add engine status file path
    excel_file, num_datasets = main(file_path1, file_path2, engine_file)
    if excel_file:
        print(f"Analysis of {num_datasets} datasets exported to {excel_file}")
    else:
        print("Failed to export analysis to Excel")